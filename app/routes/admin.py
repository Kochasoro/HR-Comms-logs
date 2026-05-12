import os
import shutil
from flask import Blueprint, current_app, jsonify, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import extract
from app.config import get_database_path
from app.models.autofill import MemoAutocomplete
from app.models.log_entry import LogEntry
from app.models.user import User
from app.models.settings import SystemSettings
from app.models.holiday import Holiday
from app.models.memo import Memo
from app.extensions import db
from datetime import date, datetime, timedelta
import pandas as pd
from io import BytesIO
from collections import defaultdict
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os

from app.services.log_service import LogService

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

@admin_bp.route("/")
@login_required
def dashboard():
    if current_user.role != "admin":
        return redirect(url_for("auth.login"))

    sort = request.args.get("sort", "latest")

    query = LogEntry.query

    if sort == "oldest":
        query = query.order_by(LogEntry.encoded_at.asc())
    else:
        query = query.order_by(LogEntry.encoded_at.desc())

    logs = query.limit(3).all()

    today = date.today()
    start_week = today - timedelta(days=today.weekday())

    week_counts = []

    for i in range(7):
        day = start_week + timedelta(days=i)

        count = Memo.query.filter(
            Memo.date == day
        ).count()

        week_counts.append(count)

    
    return render_template(
        "admin/dashboard.html",
        logs=logs,
        week_counts=week_counts,
        current_sort=sort
    )

@admin_bp.route("/settings/general", methods=["GET", "POST"])
@login_required
def settings_general():

    if current_user.role != "admin":
        return redirect(url_for("auth.login"))

    settings = SystemSettings.query.first()

    if not settings:
        settings = SystemSettings()
        db.session.add(settings)
        db.session.commit()

    if request.method == "POST":

        form_type = request.form.get("form_type")

        # GENERAL SETTINGS
        if form_type == "general":

            settings.system_name = request.form.get("system_name")

            logo_file = request.files.get("logo")

            if logo_file and logo_file.filename:

                upload_folder = os.path.join(current_app.root_path, "static", "uploads")
                os.makedirs(upload_folder, exist_ok=True)

                logo_path = os.path.join(upload_folder, "system_logo.png")
                logo_file.save(logo_path)

                settings.logo = "uploads/system_logo.png"
                
            db.session.commit()
            flash("Settings updated")

        elif form_type == "holiday":

            name = request.form.get("name")
            date_str = request.form.get("date")

            if not name or not name.strip():
                flash("Holiday name is required", "error")
                return redirect(url_for("admin.settings_general"))

            if not date_str:
                flash("Holiday date is required", "error")
                return redirect(url_for("admin.settings_general"))

            try:
                holiday_date = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                flash("Invalid date format", "error")
                return redirect(url_for("admin.settings_general"))

            repeat = "repeat_yearly" in request.form

            holiday = Holiday(
                date=holiday_date,
                name=name.strip(),
                type=request.form.get("type"),
                description=request.form.get("description"),
                repeat_yearly=repeat
            )

            db.session.add(holiday)
            db.session.commit()

            LogService.add_log(
                memo=None,
                remarks="Holiday Added",
                notes=f"{holiday.name} on {holiday.date.strftime('%Y-%m-%d')} (Repeat: {holiday.repeat_yearly})",
                user_id=current_user.id
            )

            flash("Holiday added", "success")

        return redirect(url_for("admin.settings_general"))

    holidays = Holiday.query.order_by(Holiday.date.asc()).all()
    from_offices = MemoAutocomplete.query\
        .filter_by(type="from")\
        .order_by(MemoAutocomplete.value.asc())\
        .all()

    forwarded_people = MemoAutocomplete.query\
        .filter_by(type="forwarded")\
        .order_by(MemoAutocomplete.value.asc())\
        .all()

    remarks_list = MemoAutocomplete.query\
        .filter_by(type="remarks")\
        .order_by(MemoAutocomplete.value.asc())\
        .all()

    released_people = MemoAutocomplete.query\
        .filter_by(type="released_to")\
        .order_by(MemoAutocomplete.value.asc())\
        .all()

    BASE_DIR = os.path.abspath(
        os.path.dirname(current_app.root_path)
    )

    BACKUP_DIR = os.path.join(BASE_DIR, "backups")

    STATE_FILE = os.path.join(
        BACKUP_DIR,
        "backup_state.txt"
    )

    BACKUP_A = os.path.join(
        BACKUP_DIR,
        "app_backup_A.db"
    )

    BACKUP_B = os.path.join(
        BACKUP_DIR,
        "app_backup_B.db"
    )


    backup_a_label = "Missing"
    backup_b_label = "Missing"


    if os.path.exists(STATE_FILE):

        try:

            with open(STATE_FILE, "r") as f:

                last_time, last_slot = (
                    f.read().strip().split(",")
                )

            last_backup = datetime.strptime(
                last_time,
                "%Y-%m-%d %H:%M:%S"
            )

            delta = datetime.now() - last_backup

            days = delta.days

            if days == 0:

                label = "Latest Backup"

            elif days == 1:

                label = "1 day ago"

            else:

                label = f"{days} days ago"

            if last_slot == "A":

                backup_a_label = label

                if os.path.exists(BACKUP_B):
                    backup_b_label = "Older Backup"

            else:

                backup_b_label = label

                if os.path.exists(BACKUP_A):
                    backup_a_label = "Older Backup"

        except Exception as e:

            print("Backup label error:", e)
            
    return render_template(
        "admin/settings/general.html",
        settings=settings,
        holidays=holidays,

        from_offices=from_offices,
        forwarded_people=forwarded_people,
        remarks_list=remarks_list,
        released_people=released_people,

        backup_a_label=backup_a_label,
        backup_b_label=backup_b_label
    )



@admin_bp.route("/restore-backup/<slot>", methods=["POST"])
@login_required
def restore_backup(slot):

    if current_user.role != "admin":
        return redirect(url_for("auth.login"))

    # REAL ACTIVE DATABASE
    DB_PATH = get_database_path()

    # APPDATA FOLDER
    APPDATA_DIR = os.path.dirname(DB_PATH)

    # BACKUP FOLDER
    BACKUP_DIR = os.path.join(APPDATA_DIR, "backups")

    BACKUP_A = os.path.join(
        BACKUP_DIR,
        "app_backup_A.db"
    )

    BACKUP_B = os.path.join(
        BACKUP_DIR,
        "app_backup_B.db"
    )

    backup_file = BACKUP_A if slot == "A" else BACKUP_B

    if not os.path.exists(backup_file):

        flash("Backup file not found", "error")

        return redirect(url_for("admin.settings_general"))

    try:

        # CLOSE SQLITE CONNECTIONS
        db.session.remove()
        db.engine.dispose()

        # RESTORE DATABASE
        shutil.copy2(backup_file, DB_PATH)

        flash(
            f"Backup {slot} restored successfully",
            "success"
        )

    except Exception as e:

        flash(f"Restore failed: {e}", "error")

    return redirect(url_for("admin.settings_general"))

@admin_bp.route("/delete-memos-by-date", methods=["POST"])
@login_required
def delete_memos_by_date():

    if current_user.role != "admin":
        return redirect(url_for("auth.login"))

    # 🔹 Get form values
    month_str = request.form.get("month")
    year_str = request.form.get("year")

    if not month_str or not year_str:
        flash("Month and Year are required", "error")
        return redirect(url_for("admin.settings_general"))

    # 🔹 Convert safely
    try:
        month = int(month_str)
        year = int(year_str)
    except ValueError:
        flash("Invalid month/year", "error")
        return redirect(url_for("admin.settings_general"))

    # 🔹 Create date range (THIS IS THE FIX)
    start_date = date(year, month, 1)

    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # 🔹 Query memos within the month
    memos = Memo.query.filter(
        Memo.date >= start_date,
        Memo.date < end_date
    ).all()

    print(f"Deleting {len(memos)} memos from {start_date} to {end_date}")  # debug

    if not memos:
        flash("No memos found for selected period", "error")
        return redirect(url_for("admin.settings_general"))

    try:
        count = len(memos)

        for memo in memos:
            db.session.delete(memo)

        db.session.commit()

        # 🔹 Optional: cleaner single log instead of spam
        LogService.add_log(
            memo=None,
            remarks="Batch Delete",
            notes=f"{count} memos deleted for {month}/{year}",
            user_id=current_user.id
        )

        flash(f"{count} memos deleted for {month}/{year}", "success")

    except Exception as e:
        db.session.rollback()
        print("Batch delete error:", e)
        flash("Batch delete failed", "error")

    return redirect(url_for("admin.settings_general"))

@admin_bp.route("/settings/holidays/delete/<int:id>", methods=["POST"])
@login_required
def delete_holiday(id):

    holiday = Holiday.query.get_or_404(id)

    db.session.delete(holiday)
    db.session.commit()
    LogService.add_log(
        memo=None,  
        remarks="Holiday Deleted",
        notes=f"{holiday.name} on {holiday.date.strftime('%Y-%m-%d')} (Repeat: {holiday.repeat_yearly})",
        user_id=current_user.id  
    )

    flash("Holiday deleted")

    return redirect(url_for("admin.settings_general"))
    
@admin_bp.route("/users")
@login_required
def users():

    if current_user.role != "admin":
        flash("Access denied")
        return redirect(url_for("secretary.dashboard"))

    users = User.query.all()

    return render_template("admin/user.html", users=users)
@admin_bp.route("/import")
@login_required
def import_page():  

    if current_user.role != "admin":
        flash("Access denied")
        return redirect(url_for("secretary.dashboard"))

    users = User.query.all()

    return render_template("admin/import.html", users=users)


@admin_bp.route("/import-memos", methods=["POST"])
@login_required
def import_memos():

    if current_user.role != "admin":
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    file = request.files.get("file")

    if not file:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    try:
        import uuid

        xls = pd.ExcelFile(file)

        month_map = {
            "Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
            "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12
        }

        imported = 0
        skipped = 0
        year = datetime.now().year

        # ---------- helpers ----------
        def clean(val):
            if pd.isna(val):
                return None
            return str(val).strip()

        def normalize_subject(val):
            if pd.isna(val):
                return None
            return str(val).strip().lower()

        def parse_date(val):
            if pd.isna(val):
                return None
            if isinstance(val, datetime):
                return val.date()
            try:
                return pd.to_datetime(val).date()
            except:
                return None
        # ------------------------------

        for sheet_name in xls.sheet_names:

            sheet_clean = sheet_name.split("|")[0].strip()

            if "-" not in sheet_clean:
                continue

            month_text, sheet_type = sheet_clean.split("-")

            month = month_map.get(month_text.capitalize())
            if not month:
                continue

            if "COMMUNICATION" in sheet_type.upper():
                source_type = "CM"
            elif "MEMO" in sheet_type.upper():
                source_type = "OP"
            else:
                continue

            raw_df = pd.read_excel(xls, sheet_name=sheet_name, header=None)

            # ---------- scan the starting row ----------
            start_row = None
            for i, row in raw_df.iterrows():
                val = str(row[0]).strip()

                if val.isdigit() and int(val) >= 1:
                    start_row = i
                    break

            if start_row is None:
                continue

            df = raw_df.iloc[start_row:]
            # ------------------------------------------

            for _, row in df.iterrows():

                serial_raw = clean(row[0])

                if not serial_raw:
                    skipped += 1
                    continue

                try:
                    serial = int(serial_raw)
                except:
                    skipped += 1
                    continue

                existing = Memo.query.filter_by(
                    serial_number=serial,
                    month=month,
                    year=year,
                    source_type=source_type
                ).first()

                if existing:
                    skipped += 1
                    continue

                # ---------- prepare values ----------
                date_val = parse_date(row[1])
                subject_raw = clean(row[4])
                subject_norm = normalize_subject(row[4])
                # ------------------------------------
                memo_number = None

                if source_type == "OP" and subject_raw:
                    if " - " in subject_raw:

                        split_parts = subject_raw.split(" - ", 1)

                        possible_memo = split_parts[0].strip()
                        possible_subject = split_parts[1].strip()

                        memo_number = possible_memo
                        subject_raw = possible_subject
                        subject_norm = normalize_subject(possible_subject)
                # ---------- THREADING LOGIC ----------
                thread_id = None

                if subject_norm and date_val:
                    existing_thread = Memo.query.filter(
                        Memo.subject.ilike(subject_raw.strip()),  
                        Memo.date == date_val,
                        Memo.source_type == source_type
                    ).first()

                    if existing_thread:
                        thread_id = existing_thread.thread_id

                if not thread_id:
                    thread_id = str(uuid.uuid4())
                # ------------------------------------

                memo = Memo(
                    serial_number=serial,
                    month=month,
                    year=year,

                    date=date_val,
                    from_office=clean(row[2]),
                    forwarded_by=clean(row[3]),

                    subject=subject_raw,
                    memo_number=memo_number,

                    remarks=clean(row[5]),
                    notes=clean(row[6]),

                    released_date=parse_date(row[7]),
                    released_to=clean(row[8]),

                    source_type=source_type,
                    thread_id=thread_id
                )

                db.session.add(memo)
                imported += 1

        db.session.commit()

        LogService.add_log(
            None,
            f"{current_user.username} imported memos",
            f"{imported} imported, {skipped} skipped",
            current_user.id
        )

        return jsonify({
            "success": True,
            "imported": imported,
            "skipped": skipped
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)})

@admin_bp.route("/export-memos")
@login_required
def export_memos():

    if current_user.role != "admin":
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    cm = request.args.get("cm") == "true"
    op = request.args.get("op") == "true"

    excel = request.args.get("excel") == "true"

    date_from = request.args.get("dateFrom")
    date_to = request.args.get("dateTo")

    query = Memo.query

    types = []
    if cm:
        types.append("CM")
    if op:
        types.append("OP")

    if types:
        query = query.filter(Memo.source_type.in_(types))

    if date_from:
        query = query.filter(Memo.date >= date_from)

    if date_to:
        query = query.filter(Memo.date <= date_to)

    memos = query.order_by(Memo.date).all()

    # ===== EXCEL ONLY =====
    if excel:

        output = BytesIO()

        template_path = os.path.join(
            current_app.root_path,
            "templates",
            "memo_template.xlsx"
        )

        wb = load_workbook(template_path)
        template_ws = wb.active
        template_ws.title = "TEMPLATE"

        grouped = defaultdict(list)

        for m in memos:
            date_obj = datetime.strptime(str(m.date), "%Y-%m-%d")
            month = date_obj.strftime("%b").upper()

            type_label = "COMMUNICATION" if m.source_type == "CM" else "MEMO"
            key = f"{month}-{type_label}"

            grouped[key].append(m)

        for key in sorted(grouped.keys()):
            ws = wb.copy_worksheet(template_ws)
            ws.title = key

            ws["A2"] = key
            start_row = 6

            for i, m in enumerate(grouped[key]):
                row = start_row + i

                ws.cell(row=row, column=1, value=f"{m.serial_number:04d}")
                ws.cell(row=row, column=2, value=m.date)
                ws.cell(row=row, column=3, value=m.from_office)
                ws.cell(row=row, column=4, value=m.forwarded_by)
                subject_value = (
                    f"{m.memo_number} - {m.subject}"
                    if m.source_type == "OP" and m.memo_number
                    else m.subject
                )

                ws.cell(row=row, column=5, value=subject_value)
                ws.cell(row=row, column=6, value=m.remarks)
                ws.cell(row=row, column=7, value=m.notes)
                ws.cell(row=row, column=8, value=m.released_date)
                ws.cell(row=row, column=9, value=m.released_to)

            ws.freeze_panes = "A3"

        wb.remove(template_ws)

        wb.save(output)
        output.seek(0)

        LogService.add_log(
            None,
            f"{current_user.username} exported memos as Excel",
            f"{len(memos)} memo(s) exported to XLSX",
            current_user.id
        )

        return send_file(
            output,
            as_attachment=True,
            download_name="memos_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return jsonify({"error": "Invalid export type"}), 400
    
@admin_bp.route("/users/create", methods=["POST"])
@login_required
def create_user():

    if current_user.role != "admin":
        flash("Access denied", "error")
        return redirect(url_for("secretary.dashboard"))

    username = request.form["username"].strip()
    password = request.form["password"]
    role = request.form["role"]

    # CHECK IF USER EXISTS
    existing_user = User.query.filter_by(
        username=username
    ).first()

    if existing_user:

        flash(
            f"Account '{username}' already exists.",
            "warning"
        )

        return redirect(url_for("admin.users"))

    # CREATE USER
    user = User(
        username=username,
        role=role
    )

    user.set_password(password)

    db.session.add(user)
    db.session.commit()

    LogService.add_log(
        user,
        f"{current_user.username} created a user account",
        f"Created user '{user.username}' with role '{user.role}'",
        current_user.id
    )

    flash("User created successfully", "success")

    return redirect(url_for("admin.users"))

@admin_bp.route("/users/reset/<int:id>", methods=["POST"])
@login_required
def reset_password(id):

    if current_user.role != "admin":
        flash("Access denied")
        return redirect(url_for("secretary.dashboard"))

    user = User.query.get_or_404(id)

    new_password = request.form["password"]

    user.set_password(new_password)

    db.session.commit()
    LogService.add_log(
        user,
        f"{current_user.username} reset a user's password",
        f"Password reset for '{user.username}'",
        current_user.id
    )
    flash("Password reset successfully")

    return redirect(url_for("admin.users"))

@admin_bp.route("/users/delete/<int:id>", methods=["POST"])
@login_required
def delete_user(id):

    if current_user.role != "admin":
        flash("Access denied")
        return redirect(url_for("secretary.dashboard"))

    user = User.query.get_or_404(id)

    if user.role == "superadmin":
        flash("Superadmin cannot be deleted.")
        return redirect(url_for("admin.users"))

    if user.id == current_user.id:
        flash("You cannot delete your own account.")
        return redirect(url_for("admin.users"))
    
    deleted_username = user.username
    deleted_role = user.role

    db.session.delete(user)
    db.session.commit()
    LogService.add_log(
        None,
        f"{current_user.username} deleted a user account",
        f"Deleted user '{deleted_username}' with role '{deleted_role}'",
        current_user.id
    )
    flash("User deleted")

    return redirect(url_for("admin.users"))



@admin_bp.route("/logs")
@login_required
def logs_page():

    page = request.args.get("page", 1, type=int)

    logs = LogEntry.query.order_by(
        LogEntry.encoded_at.desc()
    ).paginate(
        page=page,
        per_page=100,
        error_out=False
    )

    return render_template(
        "admin/logs.html",
        logs=logs
    )